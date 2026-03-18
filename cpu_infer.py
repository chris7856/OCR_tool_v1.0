# -*- coding: utf-8 -*-

# 作用：
# 1. 加载 PaddleOCR-VL 模型（CPU 模式）
# 2. 对输入图片/文件执行 OCR 推理
# 3. 将结果保存为 JSON / Markdown
# 4. 进一步将 Markdown 中的表格转换为 Excel
# 5. 可作为命令行脚本独立运行，也可被 GUI 子进程调用

import os
import sys
import time
import json
import logging
import traceback
import platform
import argparse
import warnings
import re
from contextlib import redirect_stdout, redirect_stderr

# HTML 表格解析
from bs4 import BeautifulSoup

# Excel 写出
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side


# ================= 控制台编码兼容处理 =================
# 在部分 Windows 环境下，控制台默认编码不是 utf-8，
# 这里尽量把 stdout/stderr 调整为 utf-8，避免中文打印报错
try:
    if hasattr(sys.stdout, "reconfigure"):
        sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    if hasattr(sys.stderr, "reconfigure"):
        sys.stderr.reconfigure(encoding="utf-8", errors="replace")
except Exception:
    pass


# ================= 依赖元数据补丁（必须最先执行） =================
# 某些第三方库在运行时会调用 importlib.metadata.version() 查询包版本。
# 但在 PyInstaller 打包环境中，有时 metadata 不完整，可能导致运行时报错。
# 因此这里做一个“版本查询兜底补丁”。
import importlib.metadata as _meta

_orig_version = _meta.version

# 为常见依赖提供兜底版本号
_SHIM = {
    "pypdfium2": "4.0.0",
    "opencv-contrib-python": "4.10.0.82",
    "opencv_contrib_python": "4.10.0.82",
    "Jinja2": "3.1.6",
    "jinja2": "3.1.6",
    "numpy": "2.0.2",
    "safetensors": "0.7.0",
    "sentencepiece": "0.2.1",
    "paddleocr": "3.4.0",
    "paddlepaddle": "3.3.0",
    "paddlex": "3.4.2",
}


def _norm_dist_name(name: str) -> str:
    """
    统一包名格式：
    - 去除首尾空格
    - 转小写
    - 下划线替换为中划线
    便于匹配 importlib.metadata 的包名查询结果
    """
    return str(name).strip().lower().replace("_", "-")


_SHIM_NORM = {_norm_dist_name(k): v for k, v in _SHIM.items()}


def _patched_version(name):
    """
    重写 importlib.metadata.version：
    若命中兜底列表，则直接返回预设版本号；
    否则调用原始 version()。
    """
    key = _norm_dist_name(name)
    if key in _SHIM_NORM:
        return _SHIM_NORM[key]
    return _orig_version(name)


# 打上补丁
_meta.version = _patched_version


# ================= 环境变量 =================
# 关闭一些不必要的检查/警告，减少控制台噪音
os.environ["PADDLE_PDX_DISABLE_MODEL_SOURCE_CHECK"] = "True"
os.environ["PADDLE_NO_COMPILE_CHECK"] = "True"
os.environ["PYTHONWARNINGS"] = "ignore"

warnings.filterwarnings("ignore")

# 把日志等级整体调低，减少第三方库输出
logging.getLogger().setLevel(logging.ERROR)
logging.getLogger("paddle").setLevel(logging.ERROR)
logging.getLogger("paddleocr").setLevel(logging.ERROR)
logging.getLogger("paddlex").setLevel(logging.ERROR)


# ================= 基础路径 =================
def get_base_path():
    """
    获取程序运行根目录。

    规则：
    - 如果是 PyInstaller 打包后的 EXE，取 exe 所在目录
    - 如果是普通 Python 脚本运行，取当前 py 文件所在目录
    """
    if getattr(sys, "frozen", False):
        return os.path.dirname(sys.executable)
    return os.path.dirname(os.path.abspath(__file__))


# 程序基础目录
BASE_DIR = get_base_path()

# 多模态 OCR 主模型目录
VL_MODEL_DIR = os.path.join(BASE_DIR, "model_file", "PaddleOCR-VL-1.5-0.9B")

# 版面检测模型目录
LAYOUT_MODEL_PATH = os.path.join(BASE_DIR, "model_file", "PP-DocLayoutV3")

# 默认测试图片路径
TEST_IMG_PATH = os.path.join(BASE_DIR, "test_img", "p1.jpg")

# 输出目录
OUTPUT_DIR = os.path.join(BASE_DIR, "output")
os.makedirs(OUTPUT_DIR, exist_ok=True)


def patch_sys_path():
    """
    将一些关键目录追加到 sys.path 前部，便于模块/资源查找。

    对打包环境和某些相对路径依赖场景比较有帮助。
    """
    extra_paths = [
        BASE_DIR,
        os.path.join(BASE_DIR, "model_file"),
        VL_MODEL_DIR,
        LAYOUT_MODEL_PATH,
    ]
    for p in extra_paths:
        if os.path.exists(p) and p not in sys.path:
            sys.path.insert(0, p)


patch_sys_path()


def log(msg):
    """
    安全打印日志信息。

    这里额外处理了编码异常，尽量避免 Windows 控制台打印中文时报错。
    """
    msg = str(msg)
    try:
        print(msg, flush=True)
    except UnicodeEncodeError:
        try:
            safe_msg = msg.encode("gbk", errors="ignore").decode("gbk", errors="ignore")
            print(safe_msg, flush=True)
        except Exception:
            try:
                print(repr(msg), flush=True)
            except Exception:
                pass


def print_header():
    """
    打印运行时环境信息，便于调试：
    - 当前是脚本模式还是 EXE 模式
    - Python 版本
    - 平台信息
    - 关键目录是否存在
    """
    log("=" * 80)
    log("[运行模式] " + ("EXE打包版" if getattr(sys, "frozen", False) else "Python脚本版"))
    log(f"[Python版本] {sys.version}")
    log(f"[平台] {platform.platform()}")
    log(f"[sys.executable] {sys.executable}")
    log(f"[BASE_DIR] {BASE_DIR}")
    log(f"[VL模型目录] {VL_MODEL_DIR} {os.path.isdir(VL_MODEL_DIR)}")
    log(f"[版面模型目录] {LAYOUT_MODEL_PATH} {os.path.isdir(LAYOUT_MODEL_PATH)}")
    log(f"[OUTPUT_DIR] {OUTPUT_DIR}")
    log("=" * 80)


def validate_environment():
    """
    检查运行环境是否完整：
    - 模型目录是否存在
    - 输出目录是否可创建

    返回：
    - 空列表：表示环境正常
    - 非空列表：表示存在问题
    """
    problems = []

    if not os.path.isdir(VL_MODEL_DIR):
        problems.append(f"VL模型目录不存在: {VL_MODEL_DIR}")

    if not os.path.isdir(LAYOUT_MODEL_PATH):
        problems.append(f"版面模型目录不存在: {LAYOUT_MODEL_PATH}")

    try:
        os.makedirs(OUTPUT_DIR, exist_ok=True)
    except Exception as e:
        problems.append(f"输出目录无法创建: {OUTPUT_DIR} -> {e}")

    return problems


class _SilentWriter:
    """
    一个“空输出流”对象。

    用于配合 redirect_stdout / redirect_stderr，
    屏蔽 Paddle / OCR 过程中的大量控制台输出。
    """
    def write(self, _):
        return 0

    def flush(self):
        return None


# ================= Markdown / HTML Table -> Excel =================
def _split_md_row(line: str):
    """
    将 Markdown 管道表格的一行拆分为单元格列表。

    例如：
    | A | B |
    -> ["A", "B"]
    """
    line = line.strip()
    if not line:
        return []

    if line.startswith("|"):
        line = line[1:]
    if line.endswith("|"):
        line = line[:-1]

    return [cell.strip() for cell in line.split("|")]


def _is_md_separator_row(cells):
    """
    判断一行是否为 Markdown 表头分隔行。

    例如：
    | --- | :---: | ---: |
    """
    if not cells:
        return False

    for c in cells:
        c = c.strip()
        if not c:
            return False
        if not re.fullmatch(r"[:\- ]+", c):
            return False
        if "-" not in c:
            return False
    return True


def parse_markdown_pipe_tables(md_text: str):
    """
    解析标准 Markdown 管道表格，例如：

    | A | B |
    | - | - |
    | 1 | 2 |

    返回格式：
    [
        [
            ["A", "B"],
            ["1", "2"]
        ],
        ...
    ]
    """
    lines = md_text.splitlines()
    tables = []
    i = 0
    n = len(lines)

    while i < n:
        line = lines[i].rstrip()

        if "|" in line:
            header_cells = _split_md_row(line)

            if i + 1 < n:
                sep_line = lines[i + 1].rstrip()
                sep_cells = _split_md_row(sep_line)

                if header_cells and _is_md_separator_row(sep_cells):
                    table = [header_cells]
                    i += 2

                    while i < n:
                        row_line = lines[i].rstrip()
                        if "|" not in row_line or not row_line.strip():
                            break
                        row_cells = _split_md_row(row_line)
                        table.append(row_cells)
                        i += 1

                    tables.append(table)
                    continue

        i += 1

    return tables


def parse_html_tables_from_markdown(md_text: str):
    """
    解析 Markdown 文本中嵌入的 HTML <table> 表格。

    返回格式：
    [
        [
            ["单元格1", "单元格2"],
            ["单元格3", "单元格4"]
        ],
        ...
    ]
    """
    soup = BeautifulSoup(md_text, "html.parser")
    html_tables = []

    for table in soup.find_all("table"):
        rows = []
        tr_list = table.find_all("tr")
        for tr in tr_list:
            cells = tr.find_all(["th", "td"])
            row = []
            for cell in cells:
                text = cell.get_text(separator=" ", strip=True)
                row.append(text)
            if row:
                rows.append(row)
        if rows:
            html_tables.append(rows)

    return html_tables


def _apply_excel_style(ws):
    """
    为工作表设置基础样式：
    - 居中对齐
    - 自动换行
    - 单元格边框
    - 表头加粗并设置底色
    """
    thin = Side(style="thin", color="CCCCCC")

    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

    if ws.max_row >= 1:
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="D9EAF7")


def _auto_adjust_column_width(ws, min_width=10, max_width=40):
    """
    自动调整列宽。

    根据每列最长文本长度估算宽度，并设置最小/最大宽度限制，
    避免列太窄看不清，也避免列过宽影响布局。
    """
    for col_cells in ws.columns:
        max_len = 0
        col_letter = col_cells[0].column_letter
        for cell in col_cells:
            v = "" if cell.value is None else str(cell.value)
            lines = v.splitlines() if v else [""]
            local_max = max((len(x) for x in lines), default=0)
            if local_max > max_len:
                max_len = local_max
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, min_width), max_width)


def _write_table_to_sheet(ws, table):
    """
    将二维表格数据写入指定工作表，并应用样式。
    """
    if not table:
        return

    max_cols = max(len(row) for row in table)

    for row in table:
        padded = row + [""] * (max_cols - len(row))
        ws.append(padded)

    _apply_excel_style(ws)
    _auto_adjust_column_width(ws)
    ws.freeze_panes = "A2"


def markdown_to_excel(md_path, excel_path):
    """
    将 Markdown 文件转换为 Excel 文件。

    处理优先级：
    1. 优先解析 Markdown 中的 HTML table
    2. 若没有 HTML table，再解析标准 Markdown 管道表格
    3. 若仍未识别到表格，则按全文逐行写入一个工作表
    """
    with open(md_path, "r", encoding="utf-8") as f:
        md_text = f.read()

    html_tables = parse_html_tables_from_markdown(md_text)
    pipe_tables = parse_markdown_pipe_tables(md_text)

    wb = Workbook()
    default_ws = wb.active

    if html_tables:
        # 如果识别到 HTML 表格，则每个表格写入一个 sheet
        wb.remove(default_ws)
        for idx, table in enumerate(html_tables, start=1):
            ws = wb.create_sheet(title=f"Table_{idx}")
            _write_table_to_sheet(ws, table)

    elif pipe_tables:
        # 如果识别到 Markdown 管道表格，则每个表格写入一个 sheet
        wb.remove(default_ws)
        for idx, table in enumerate(pipe_tables, start=1):
            ws = wb.create_sheet(title=f"Table_{idx}")
            _write_table_to_sheet(ws, table)

    else:
        # 如果没有表格，则按普通文本逐行写入
        default_ws.title = "Markdown"
        default_ws.append(["内容"])
        for line in md_text.splitlines():
            default_ws.append([line])

        default_ws.column_dimensions["A"].width = 120
        default_ws["A1"].font = Font(bold=True)
        default_ws["A1"].fill = PatternFill("solid", fgColor="D9EAF7")
        default_ws.freeze_panes = "A2"

    # 确保输出目录存在
    os.makedirs(os.path.dirname(os.path.abspath(excel_path)), exist_ok=True)
    wb.save(excel_path)


# ================= 推理相关 =================
def load_pipeline():
    """
    加载 PaddleOCR-VL 推理管线。

    返回：
    - pipeline: 已加载好的模型对象
    - load_cost: 模型加载耗时（秒）
    """
    log("开始加载模型（CPU首次加载较慢）...")
    start_load = time.time()

    silent = _SilentWriter()

    # 屏蔽模型加载时的冗余输出
    with redirect_stdout(silent), redirect_stderr(silent):
        from paddleocr import PaddleOCRVL
        pipeline = PaddleOCRVL(
            vl_rec_model_dir=VL_MODEL_DIR,
            layout_detection_model_dir=LAYOUT_MODEL_PATH,
            device="cpu",
            precision="fp32"
        )

    load_cost = round(time.time() - start_load, 2)
    log(f"模型加载完成，耗时：{load_cost} 秒")
    return pipeline, load_cost


def infer_one_file(pipeline, input_path, output_dir):
    """
    对单个文件执行推理，并保存结果文件。

    输出内容可能包括：
    - JSON
    - Markdown
    - Excel（由 Markdown 二次转换而来）
    """
    input_path = os.path.abspath(input_path)
    if not os.path.isfile(input_path):
        raise FileNotFoundError(f"输入文件不存在: {input_path}")

    os.makedirs(output_dir, exist_ok=True)

    log("-" * 80)
    log(f"开始推理文件: {os.path.basename(input_path)}")
    start_infer = time.time()

    silent = _SilentWriter()

    # 屏蔽推理期间第三方库输出
    with redirect_stdout(silent), redirect_stderr(silent):
        output = pipeline.predict(input_path)

    infer_cost = round(time.time() - start_infer, 2)
    log(f"推理完成，耗时：{infer_cost} 秒")

    # 生成输出文件名
    img_name = os.path.splitext(os.path.basename(input_path))[0]
    json_path = os.path.join(output_dir, f"{img_name}_res.json")
    md_path = os.path.join(output_dir, f"{img_name}_res.md")
    excel_path = os.path.join(output_dir, f"{img_name}_res.xlsx")

    result_count = 0
    generated_json_paths = []
    generated_md_paths = []
    generated_excel_paths = []

    if output:
        output_list = list(output)
        total = len(output_list)

        for idx, res in enumerate(output_list, start=1):
            result_count += 1

            # 如果只有一个结果，使用固定名称；
            # 如果有多个结果，则加序号避免覆盖
            save_json_path = json_path if total == 1 else os.path.join(output_dir, f"{img_name}_res_{idx}.json")
            save_md_path = md_path if total == 1 else os.path.join(output_dir, f"{img_name}_res_{idx}.md")
            save_excel_path = excel_path if total == 1 else os.path.join(output_dir, f"{img_name}_res_{idx}.xlsx")

            # 保存 JSON
            try:
                with redirect_stdout(silent), redirect_stderr(silent):
                    res.save_to_json(save_json_path)
                generated_json_paths.append(save_json_path)
            except Exception as e:
                log(f"[警告] save_to_json 失败: {e}")

            # 保存 Markdown
            try:
                with redirect_stdout(silent), redirect_stderr(silent):
                    res.save_to_markdown(save_md_path)
                generated_md_paths.append(save_md_path)
            except Exception as e:
                log(f"[警告] save_to_markdown 失败: {e}")

            # 将 Markdown 转换为 Excel
            try:
                if os.path.isfile(save_md_path):
                    markdown_to_excel(save_md_path, save_excel_path)
                    generated_excel_paths.append(save_excel_path)
                else:
                    log(f"[警告] Excel 生成失败，markdown不存在: {save_md_path}")
            except Exception as e:
                log(f"[警告] markdown_to_excel 失败: {e}")
    else:
        log("[提示] 未识别到任何内容")

    return {
        "ok": True,
        "input_path": input_path,
        "json_path": json_path,
        "md_path": md_path,
        "excel_path": excel_path,
        "json_paths": generated_json_paths,
        "md_paths": generated_md_paths,
        "excel_paths": generated_excel_paths,
        "result_count": result_count,
        "infer_cost": infer_cost,
    }


def write_result_json(result, result_file):
    """
    将整体执行结果写入 JSON 文件，便于 GUI 或其他程序读取。
    """
    os.makedirs(os.path.dirname(os.path.abspath(result_file)), exist_ok=True)
    with open(result_file, "w", encoding="utf-8") as f:
        json.dump(result, f, ensure_ascii=False, indent=2)


def run_batch(inputs, output_dir):
    """
    批量处理多个输入文件。

    流程：
    1. 检查环境
    2. 加载模型（只加载一次）
    3. 逐个文件执行推理
    4. 汇总成功/失败情况
    """
    problems = validate_environment()
    if problems:
        raise RuntimeError("环境检查失败:\n" + "\n".join(problems))

    if not inputs:
        raise RuntimeError("未提供输入文件")

    start_all = time.time()

    # 模型只加载一次，后续复用
    pipeline, load_cost = load_pipeline()

    items = []
    ok_count = 0
    fail_count = 0

    for path in inputs:
        try:
            item = infer_one_file(pipeline, path, output_dir)
            ok_count += 1
            item["ok"] = True
        except Exception as e:
            fail_count += 1
            item = {
                "ok": False,
                "input_path": os.path.abspath(path),
                "error": repr(e),
                "traceback": traceback.format_exc(),
            }
            log(f"文件处理失败: {os.path.basename(path)}")

        items.append(item)

    total_cost = round(time.time() - start_all, 2)

    result = {
        "ok": fail_count == 0,
        "load_cost": load_cost,
        "total_cost": total_cost,
        "ok_count": ok_count,
        "fail_count": fail_count,
        "items": items,
    }
    return result


def cli_main():
    """
    命令行主入口。

    支持：
    - 直接通过 --input 重复传入多个文件
    - 通过 --input_list_file 传入 txt/json 列表文件
    - 指定输出目录
    - 指定结果汇总 JSON 文件
    """
    parser = argparse.ArgumentParser()
    parser.add_argument("--input", action="append", help="输入文件路径，可重复传入多个")
    parser.add_argument("--input_list_file", default="", help="包含多个输入文件路径的txt/json文件")
    parser.add_argument("--output_dir", default=OUTPUT_DIR, help="输出目录")
    parser.add_argument("--result_file", default="", help="结果json文件路径")
    parser.add_argument("--debug", action="store_true", help="打印调试头信息")
    args = parser.parse_args()

    try:
        if args.debug:
            print_header()

        inputs = []

        # 处理命令行直接传入的多个 --input
        if args.input:
            inputs.extend(args.input)

        # 处理列表文件
        if args.input_list_file:
            list_file = os.path.abspath(args.input_list_file)
            if not os.path.isfile(list_file):
                raise FileNotFoundError(f"输入列表文件不存在: {list_file}")

            # 如果是 JSON，则要求内容为数组
            if list_file.lower().endswith(".json"):
                with open(list_file, "r", encoding="utf-8") as f:
                    data = json.load(f)
                if not isinstance(data, list):
                    raise RuntimeError("JSON 输入列表文件必须是数组")
                inputs.extend([str(x) for x in data])
            else:
                # 否则按普通 txt 文件逐行读取
                with open(list_file, "r", encoding="utf-8") as f:
                    for line in f:
                        line = line.strip()
                        if line:
                            inputs.append(line)

        # 如果没有指定输入，则尝试使用默认测试图片
        if not inputs:
            if not os.path.isfile(TEST_IMG_PATH):
                raise FileNotFoundError(f"默认图片不存在: {TEST_IMG_PATH}")
            inputs = [TEST_IMG_PATH]

        result = run_batch(inputs, args.output_dir)

        # 如指定结果文件，则写出汇总 JSON
        if args.result_file:
            write_result_json(result, args.result_file)

        # 全部成功返回 0，否则返回 1
        return 0 if result["fail_count"] == 0 else 1

    except Exception as e:
        err_tb = traceback.format_exc()

        log("")
        log("运行失败: " + repr(e))
        log("")
        log("[完整异常堆栈]")
        log(err_tb)

        # 保存错误日志，便于排查打包运行异常
        final_output = args.output_dir or OUTPUT_DIR
        os.makedirs(final_output, exist_ok=True)
        err_log = os.path.join(final_output, "error_log.txt")
        try:
            with open(err_log, "a", encoding="utf-8") as f:
                f.write("\n" + "=" * 80 + "\n")
                f.write(f"[错误] {repr(e)}\n\n")
                f.write(err_tb)
                f.write("\n")
            log(f"[错误日志已保存] {err_log}")
        except Exception as log_e:
            log(f"[写错误日志失败] {log_e}")

        # 如果调用方要求输出结果 JSON，则将异常也写进去
        if args.result_file:
            try:
                write_result_json({
                    "ok": False,
                    "error": repr(e),
                    "traceback": err_tb,
                }, args.result_file)
            except Exception:
                pass

        return 1


if __name__ == "__main__":
    # Windows + PyInstaller 打包环境下，
    # 如果涉及 multiprocessing，需要调用 freeze_support()
    if sys.platform == "win32" and getattr(sys, "frozen", False):
        import multiprocessing
        multiprocessing.freeze_support()

    code = cli_main()

    # 如果是双击 EXE 且没有命令行参数，执行完后暂停一下，
    # 避免窗口一闪而过，方便查看日志
    if getattr(sys, "frozen", False) and len(sys.argv) == 1:
        try:
            input("\n运行完成！按回车退出...")
        except Exception:
            pass

    sys.exit(code)
